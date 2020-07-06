import logging
import os
import sys
import time
import openpyxl
import psutil
from PyQt5 import QtCore, QtGui, QtWidgets, Qt
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import platform
import runpy

platform = platform.system()


if platform == "win32" or platform == "Windows":
    from configs.Windows import Paths 

elif platform == "Linux" or platform == "Linux2":
    from configs.Linux import Paths


TIME_LIMIT = 100


# Functions for main window buttons signal
def translations():
    # Translations box
    Dialog.show()


def automation():
    # Automation box
    Dialog2.show()


def wrong_domain():
    # Tests box
    Domain_warning.show()


def script_finished():
    # Tests box
    Script_finished_popup.show()


def kill_driver():
    # End chromedriver and chrome ghost processes
    PROCNAME = "chromedriver.exe"
    PROCNAME2 = "chrome.exe"

    for proc in psutil.process_iter():
        # check whether the process name matches
        # if proc.name() == PROCNAME2:
        #     try:
        #         proc.kill()
        #
        #     except psutil.NoSuchProcess:
        #         break

        if proc.name() == PROCNAME:
            try:
                proc.kill()

            except psutil.NoSuchProcess:
                break


# noinspection PyShadowingNames,PyUnresolvedReferences
class Ui_Dialog:
    # Translation dialog
    # noinspection PyAttributeOutsideInit
    def setupUi(self, Dialog):
        self.threadpool = QThreadPool()

        def change_translation():
            # Get values from import boxes
            domain = str(self.comboBox.currentText())

            key_name = self.textEdit_1.toPlainText()
            self.textEdit_1.setPlainText(key_name)

            old_pl = self.textEdit_2.toPlainText()
            self.textEdit_2.setPlainText(old_pl)

            old_en = self.textEdit_3.toPlainText()
            self.textEdit_3.setPlainText(old_en)

            new_pl = self.textEdit_4.toPlainText()
            self.textEdit_4.setPlainText(new_pl)

            new_en = self.textEdit_5.toPlainText()
            self.textEdit_5.setPlainText(new_en)

            # Save values from input boxes to xlsx file for script use
            WB = openpyxl.load_workbook(Paths.XLSX)
            SHEET = WB.active
            ROWNUM = SHEET.max_row + 1

            """Inputs log values into Excel file ready to report"""
            string_to_display = key_name
            logging.info(string_to_display)

            string_to_display = new_pl
            logging.info(string_to_display)

            string_to_display = new_en
            logging.info(string_to_display)

            SHEET.cell(row=ROWNUM, column=2).value = domain.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=3).value = key_name.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=4).value = new_pl.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=5).value = new_en.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=6).value = old_pl.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=7).value = old_en.lstrip().rstrip()

            time_now = time.strftime("%x")
            SHEET.cell(row=ROWNUM, column=1).value = time_now

            WB.save(Paths.XLSX)
            time.sleep(5)
            runpy.run_path(Paths.Translations_change_keys)
            self.textEdit_5.setPlainText(new_en)
            script_finished()

        def summernote():
            # Get values from import boxes
            domain = str(self.comboBox.currentText())

            key_name = self.textEdit_6.toPlainText()
            self.textEdit_6.setPlainText(key_name)

            new_pl = self.textEdit_7.toPlainText()
            self.textEdit_7.setPlainText(new_pl)

            new_en = self.textEdit_8.toPlainText()
            self.textEdit_8.setPlainText(new_en)

            # Save values from input boxes to xlsx file for script use
            WB = openpyxl.load_workbook(Paths.XLSX)
            SHEET = WB.active
            ROWNUM = SHEET.max_row + 1

            """Inputs log values into Excel file ready to report"""
            string_to_display = key_name
            logging.info(string_to_display)

            string_to_display = new_pl
            logging.info(string_to_display)

            string_to_display = new_en
            logging.info(string_to_display)

            SHEET.cell(row=ROWNUM, column=2).value = domain.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=3).value = key_name.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=4).value = new_pl.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=5).value = new_en.lstrip().rstrip()

            time_now = time.strftime("%x")
            SHEET.cell(row=ROWNUM, column=1).value = time_now

            WB.save(Paths.XLSX)
            time.sleep(5)
            runpy.run_path(Paths.Translations_summernote)
            self.textEdit_8.setPlainText(new_en)
            script_finished()

        def add_new():
            # Get values from import boxes
            domain = str(self.comboBox.currentText())

            key_name = self.textEdit_9.toPlainText()
            self.textEdit_9.setPlainText(key_name)

            new_pl = self.textEdit_10.toPlainText()
            self.textEdit_10.setPlainText(new_pl)

            new_en = self.textEdit_11.toPlainText()
            self.textEdit_11.setPlainText(new_en)

            # Save values from input boxes to xlsx file for script use
            WB = openpyxl.load_workbook(Paths.XLSX)
            SHEET = WB.active
            ROWNUM = SHEET.max_row + 1

            """Inputs log values into Excel file ready to report"""
            string_to_display = key_name
            logging.info(string_to_display)

            string_to_display = new_pl
            logging.info(string_to_display)

            string_to_display = new_en
            logging.info(string_to_display)

            SHEET.cell(row=ROWNUM, column=2).value = domain.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=3).value = key_name.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=4).value = new_pl.lstrip().rstrip()
            SHEET.cell(row=ROWNUM, column=5).value = new_en.lstrip().rstrip()

            time_now = time.strftime("%x")
            SHEET.cell(row=ROWNUM, column=1).value = time_now

            WB.save(Paths.XLSX)
            time.sleep(5)
            runpy.run_path(Paths.Translations_new_keys)
            script_finished()

        def get_tab():
            # 0: Change translations
            # 1: Summernote
            # 2: Add new
            tab = self.tabWidget.currentIndex()
            domain = self.comboBox.currentText()

            if domain == "--Domain--":
                wrong_domain()

            else:
                if tab == 0:
                    change_translation()

                if tab == 1:
                    summernote()

                if tab == 2:
                    add_new()

        Dialog.setWindowTitle("Translate")
        Dialog.resize(491, 375)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(Paths.favicon), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        Dialog.setWindowIcon(icon)
        Dialog.setStyleSheet("background-color: white;")

        self.buttonBox = QtWidgets.QDialogButtonBox(Dialog)
        self.buttonBox.setGeometry(QtCore.QRect(170, 340, 171, 32))
        self.buttonBox.setAutoFillBackground(False)
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.pushButton_2 = QtWidgets.QPushButton(Dialog)
        self.pushButton_2.setGeometry(QtCore.QRect(190, 340, 101, 31))
        self.pushButton_2.setText("Start")
        self.buttonBox.setCenterButtons(True)

        self.comboBox = QtWidgets.QComboBox(Dialog)
        self.comboBox.setGeometry(QtCore.QRect(150, 310, 171, 21))
        self.comboBox.setCurrentText("--Domain--")
        self.comboBox.addItem("--Domain--")
        self.comboBox.addItem("validators")
        self.comboBox.addItem("messages")
        self.comboBox.addItem("whr")
        self.comboBox.addItem("forms")

        self.pushButton = QtWidgets.QPushButton(Dialog)
        self.pushButton.setGeometry(QtCore.QRect(350, 310, 75, 23))

        self.tabWidget = QtWidgets.QTabWidget(Dialog)
        self.tabWidget.setGeometry(QtCore.QRect(0, 0, 491, 301))
        self.tabWidget.setTabPosition(QtWidgets.QTabWidget.North)
        self.tabWidget.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.tabWidget.setElideMode(QtCore.Qt.ElideNone)

        self.tab = QtWidgets.QWidget()
        self.tab_2 = QtWidgets.QWidget()
        self.tab_3 = QtWidgets.QWidget()

        # Add 3 tabs widgets to window
        self.tabWidget.addTab(self.tab, "Change translations")
        self.tabWidget.addTab(self.tab_2, "Summernote")
        self.tabWidget.addTab(self.tab_3, "Add new")

        # Add layout for input fields
        self.verticalLayoutWidget = QtWidgets.QWidget(self.tab)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(20, 10, 441, 261))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)

        self.textEdit_1 = QtWidgets.QTextEdit(self.verticalLayoutWidget)
        self.textEdit_1.setPlaceholderText("App.key name")

        self.textEdit_2 = QtWidgets.QTextEdit(self.verticalLayoutWidget)
        self.textEdit_2.setPlaceholderText("Old Polish translation name")

        self.textEdit_3 = QtWidgets.QTextEdit(self.verticalLayoutWidget)
        self.textEdit_3.setPlaceholderText("Old English translation name")

        self.textEdit_4 = QtWidgets.QTextEdit(self.verticalLayoutWidget)
        self.textEdit_4.setPlaceholderText("New Polish translation name")

        self.textEdit_5 = QtWidgets.QTextEdit(self.verticalLayoutWidget)
        self.textEdit_5.setPlaceholderText("New English translation name")

        # Add text inputs to tab#2- Change translations
        self.verticalLayout.addWidget(self.textEdit_1)
        self.verticalLayout.addWidget(self.textEdit_2)
        self.verticalLayout.addWidget(self.textEdit_3)
        self.verticalLayout.addWidget(self.textEdit_4)
        self.verticalLayout.addWidget(self.textEdit_5)

        self.verticalLayoutWidget_2 = QtWidgets.QWidget(self.tab_2)
        self.verticalLayoutWidget_2.setGeometry(QtCore.QRect(20, 10, 441, 261))

        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_2)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)

        self.textEdit_6 = QtWidgets.QTextEdit(self.verticalLayoutWidget_2)
        self.textEdit_6.setPlaceholderText("App.key name")

        self.textEdit_7 = QtWidgets.QTextEdit(self.verticalLayoutWidget_2)
        self.textEdit_7.setPlaceholderText("New Polish translation name")

        self.textEdit_8 = QtWidgets.QTextEdit(self.verticalLayoutWidget_2)
        self.textEdit_8.setPlaceholderText("New English translation name")

        # Add text inputs to tab#2- Summernote
        self.verticalLayout_2.addWidget(self.textEdit_6)
        self.verticalLayout_2.addWidget(self.textEdit_7)
        self.verticalLayout_2.addWidget(self.textEdit_8)
        self.verticalLayout_2.addWidget(self.textEdit_7)
        self.verticalLayout_2.addWidget(self.textEdit_8)

        self.verticalLayoutWidget_3 = QtWidgets.QWidget(self.tab_3)
        self.verticalLayoutWidget_3.setGeometry(QtCore.QRect(20, 10, 441, 261))

        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_3)
        self.verticalLayout_3.setContentsMargins(0, 0, 0, 0)

        self.textEdit_9 = QtWidgets.QTextEdit(self.verticalLayoutWidget_3)
        self.textEdit_9.setPlaceholderText("App.key name")

        self.textEdit_10 = QtWidgets.QTextEdit(self.verticalLayoutWidget_3)
        self.textEdit_10.setPlaceholderText("New Polish translation name")

        self.textEdit_11 = QtWidgets.QTextEdit(self.verticalLayoutWidget_3)
        self.textEdit_11.setPlaceholderText("New English translation name")

        # Add text inputs to tab#3- Add new
        self.verticalLayout_3.addWidget(self.textEdit_9)
        self.verticalLayout_3.addWidget(self.textEdit_10)
        self.verticalLayout_3.addWidget(self.textEdit_11)

        # Center text in comboBox
        self.comboBox.setEditable(True)
        line_edit = self.comboBox.lineEdit()
        line_edit.setAlignment(Qt.AlignCenter)
        line_edit.setReadOnly(True)

        # Connect signals to buttons
        self.tabWidget.setCurrentIndex(0)
        self.pushButton_2.clicked.connect(get_tab)
        self.buttonBox.rejected.connect(Dialog.reject)
        self.pushButton.setText("Clear all")
        self.pushButton.clicked.connect(self.textEdit_1.clear)
        self.pushButton.clicked.connect(self.textEdit_2.clear)
        self.pushButton.clicked.connect(self.textEdit_3.clear)
        self.pushButton.clicked.connect(self.textEdit_4.clear)
        self.pushButton.clicked.connect(self.textEdit_5.clear)
        self.pushButton.clicked.connect(self.textEdit_6.clear)
        self.pushButton.clicked.connect(self.textEdit_7.clear)
        self.pushButton.clicked.connect(self.textEdit_8.clear)
        self.pushButton.clicked.connect(self.textEdit_7.clear)
        self.pushButton.clicked.connect(self.textEdit_8.clear)
        self.pushButton.clicked.connect(self.textEdit_9.clear)
        self.pushButton.clicked.connect(self.textEdit_10.clear)
        self.pushButton.clicked.connect(self.textEdit_11.clear)

        QtCore.QMetaObject.connectSlotsByName(Dialog)
        Dialog.setTabOrder(self.comboBox, self.textEdit_2)
        Dialog.setTabOrder(self.textEdit_1, self.pushButton)
        Dialog.setTabOrder(self.textEdit_2, self.textEdit_3)
        Dialog.setTabOrder(self.textEdit_3, self.textEdit_4)
        Dialog.setTabOrder(self.textEdit_4, self.textEdit_1)


# noinspection PyShadowingNames,PyAttributeOutsideInit,PyUnresolvedReferences
class Ui_Dialog2:
    # Automation dialog
    def setupUi(self, Dialog):
        self.threadpool = QThreadPool()
        # Window setup
        Dialog.setWindowTitle("Automation")
        Dialog.resize(420, 300)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(Paths.favicon), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        Dialog.setWindowIcon(icon)
        Dialog.setStyleSheet("background-color: white;")

        label2 = QLabel(Dialog)
        pixmap = QPixmap(Paths.graphics)
        label2.setPixmap(pixmap)

        def result():
            text = str(self.comboBox.currentText())

            if text == 'Payrolls':
                import Other.Payrolls
                script_finished()

            if text == 'Copy Translations':
                import Translations.Copy_translations
                script_finished()

            if text == 'Change passwords':
                import Other.Change_passwords
                script_finished()

        self.comboBox = QtWidgets.QComboBox(Dialog)
        self.comboBox.setGeometry(QtCore.QRect(130, 210, 160, 21))
        self.comboBox.addItem("-- Script --")
        self.comboBox.addItem("Change passwords")
        self.comboBox.addItem("Payrolls")
        self.comboBox.addItem("Copy Translations")

        self.comboBox.setEditable(True)
        line_edit = self.comboBox.lineEdit()
        line_edit.setAlignment(Qt.AlignCenter)
        line_edit.setReadOnly(True)

        self.pushButton = QtWidgets.QPushButton(Dialog)
        self.pushButton.setGeometry(QtCore.QRect(160, 250, 91, 31))
        self.pushButton.setText("Start")
        self.pushButton.clicked.connect(result)

        QtCore.QMetaObject.connectSlotsByName(Dialog)


# noinspection PyAttributeOutsideInit,PyShadowingNames,PyTypeChecker
class Warning_pop:
    def setupUi(self, Dialog):
        self.threadpool = QThreadPool()
        # Window setup
        Dialog.setWindowTitle("Warning!")
        Dialog.resize(372, 194)
        Dialog.setAutoFillBackground(True)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(Paths.favicon), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        Dialog.setWindowIcon(icon)
        Dialog.setStyleSheet("background-color: white;")

        self.buttonBox = QtWidgets.QDialogButtonBox(Dialog)
        self.buttonBox.setGeometry(QtCore.QRect(30, 120, 241, 32))
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel | QtWidgets.QDialogButtonBox.Ok)

        self.textEdit = QtWidgets.QTextEdit(Dialog)
        self.textEdit.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.textEdit.setGeometry(QtCore.QRect(40, 20, 291, 91))
        self.textEdit.setReadOnly(True)
        self.textEdit.setPlainText("You've chosen unsupported script, try another one")

        # Buttons in popup
        self.buttonBox.accepted.connect(Dialog.accept)
        self.buttonBox.rejected.connect(Dialog.reject)
        QtCore.QMetaObject.connectSlotsByName(Dialog)


# noinspection PyAttributeOutsideInit,PyShadowingNames,PyTypeChecker,PyTypeChecker
class Domain_warning:
    def setupUi(self, Dialog):
        # Window setup
        Dialog.setWindowTitle("Warning!")
        Dialog.resize(372, 194)
        Dialog.setAutoFillBackground(True)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(Paths.favicon), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        Dialog.setWindowIcon(icon)
        Dialog.setStyleSheet("background-color: white;")

        self.buttonBox = QtWidgets.QDialogButtonBox(Dialog)
        self.buttonBox.setGeometry(QtCore.QRect(30, 120, 241, 32))
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel | QtWidgets.QDialogButtonBox.Ok)

        self.textEdit = QtWidgets.QTextEdit(Dialog)
        self.textEdit.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.textEdit.setGeometry(QtCore.QRect(40, 20, 291, 91))
        self.textEdit.setReadOnly(True)
        self.textEdit.setPlainText("You didn't choose correct domain, try again!")

        # Buttons in popup
        self.buttonBox.accepted.connect(Dialog.accept)
        self.buttonBox.rejected.connect(Dialog.reject)
        QtCore.QMetaObject.connectSlotsByName(Dialog)


class Script_finished_dialog(object):
    def setupUi(self, Dialog):
        def Close_window():
            Dialog.close()

        Dialog.setObjectName("Success")
        Dialog.resize(377, 188)
        Dialog.setStyleSheet("background-color: white;")

        self.pushButton = QPushButton(Dialog)
        self.textEdit = QtWidgets.QTextEdit(Dialog)

        self.pushButton.setText("OK")
        self.pushButton.setGeometry(QRect(140, 120, 89, 25))
        self.pushButton.clicked.connect(Close_window)
        self.textEdit.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.textEdit.setGeometry(QtCore.QRect(40, 20, 291, 91))
        self.textEdit.setPlainText("Script completed successfully")
        self.textEdit.setAlignment(Qt.AlignCenter)
        self.textEdit.setReadOnly(True)
        QMetaObject.connectSlotsByName(Dialog)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    w = QWidget()
    w.resize(420, 245)
    w.setStyleSheet("background-color: white;")
    w.setWindowTitle('Extranet automation')
    icon = QtGui.QIcon()
    icon.addPixmap(QtGui.QPixmap(Paths.favicon), QtGui.QIcon.Normal, QtGui.QIcon.Off)
    app.setWindowIcon(icon)

    label = QLabel(w)
    pixmap = QPixmap(Paths.graphics)
    label.setPixmap(pixmap)
    label.move(0, 0)

    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)

    Dialog2 = QtWidgets.QDialog()
    ui2 = Ui_Dialog2()
    ui2.setupUi(Dialog2)

    Warning_popup = QtWidgets.QDialog()
    ui4 = Warning_pop()
    ui4.setupUi(Warning_popup)

    Domain_warning = QtWidgets.QDialog()
    ui5 = Warning_pop()
    ui5.setupUi(Domain_warning)

    Script_finished_popup = QtWidgets.QDialog()
    ui6 = Script_finished_dialog()
    ui6.setupUi(Script_finished_popup)

    # Add buttons to main window- select modules
    btn = QPushButton(w)
    btn.setText('Automation')
    btn.move(50, 192)
    btn.resize(100, 50)
    btn.show()
    btn.clicked.connect(automation)

    btn2 = QPushButton(w)
    btn2.setText('Translation')
    btn2.move(150, 192)
    btn2.resize(100, 50)
    btn2.show()
    btn2.clicked.connect(translations)

    btn4 = QPushButton(w)
    btn4.setText('Kill driver')
    btn4.move(250, 192)
    btn4.resize(100, 50)
    btn4.show()
    btn4.clicked.connect(kill_driver)

    w.show()
    sys.exit(app.exec_())
