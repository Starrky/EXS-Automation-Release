"""Keys and translations changed Prod/ Staging """
# !/usr/bin/python3.7.2
# coding=utf-8
import os
import sys
import time
from urllib.parse import urljoin
import openpyxl
import psutil
import selenium
import webdriver_manager
from playsound import playsound
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import platform

sys.path.append(
    os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)))

from configs import Accounts
from configs import Sites

platform = platform.system()

if platform == "win32" or platform == "Windows":
    from configs.Windows import Paths

elif platform == "Linux" or platform == "Linux2":
    from configs.Linux import Paths

# Driver settings
DRIVER = webdriver.Chrome(webdriver_manager.chrome.ChromeDriverManager().install())


def kill_driver():
    # End chromedriver and chrome ghost processes
    PROCNAME = "chromedriver.exe"
    PROCNAME2 = "chrome.exe"

    for proc in psutil.process_iter():
        # check whether the process name matches
        if proc.name() == PROCNAME:
            try:
                proc.kill()

            except psutil.NoSuchProcess:
                break


websites_list = [Sites.Prod, Sites.Staging]
logins_list = [Accounts.Prod_user_1, Accounts.Staging_user_2]
passwords_list = [Accounts.Prod_Password_1, Accounts.Staging_Password_2]
copy_list = [Sites.Prod_Copy, Sites.Staging_Copy]
logout_list = [Sites.Staging_logout, Sites.Prod_logout]

# Grab text from launcher
currentdir = os.getcwd()
WB = openpyxl.load_workbook(f"{currentdir}/List.xlsx")
SHEET = WB.active
ROWNUM = SHEET.max_row

Keys_List = []
Domains = []
New_pl = []
New_en = []

colDict = {1: Domains, 2: Keys_List, 5: New_pl, 6: New_en}

for colNum, obj in colDict.items():
    for row in SHEET.iter_rows(min_row=2, min_col=colNum, max_col=colNum, max_row=ROWNUM):
        for cell in row:
            obj.append(cell.value)


def create_new_pl():
    """Create New polish translation for the key"""
    fa_plus = DRIVER.find_element_by_css_selector(".fa-plus")
    fa_plus.click()
    for key, new_pl in zip(Keys_List, New_pl):
        create_key = DRIVER.find_element_by_id("create-key")
        create_message = DRIVER.find_element_by_id("create-message")
        create_key.click()
        create_key.clear()
        create_key.send_keys(str(key))
        create_message.click()
        create_message.clear()
        create_message.send_keys(str(new_pl))
        create_message.send_keys(Keys.RETURN)
        time.sleep(4)


def create_new(): 
    """Create New english translation for the key(all languages, but polish and ref)"""
    fa_plus = DRIVER.find_element_by_css_selector(".fa-plus")
    fa_plus.click()

    for key, new_en in zip(Keys_List, New_en):
        create_key = DRIVER.find_element_by_id("create-key")
        create_message = DRIVER.find_element_by_id("create-message")
        create_key.click()
        create_key.clear()
        create_key.send_keys(str(key))
        create_message.click()
        create_message.clear()
        create_message.send_keys(str(new_en))
        create_message.send_keys(Keys.RETURN)
        time.sleep(4)


def create_new_ref():
    """Create New REF translation for the key"""
    fa_plus = DRIVER.find_element_by_css_selector(".fa-plus")
    fa_plus.click()
    for key in Keys_List:
        create_key = DRIVER.find_element_by_id("create-key")
        create_message = DRIVER.find_element_by_id("create-message")
        create_key.click()
        create_key.clear()
        create_key.send_keys(str(key))
        create_message.click()
        create_message.clear()
        create_message.send_keys(str(key))
        create_message.send_keys(Keys.RETURN)
        time.sleep(4)

def translate():
    """Actual translation execution code- takes domain from excel file and takes credentials from Accounts
     file, then executes script on both staging and prod site"""
    DRIVER.start_client()
    for site, login, password, domain in zip(websites_list, logins_list, passwords_list, Domains):
        while True:
            try:
                DRIVER.get(site)
                Password_field = DRIVER.find_element_by_id("password")
                username = DRIVER.find_element_by_id("username")
                username.click()
                username.clear()
                username.send_keys(login)
                Password_field.click()
                Password_field.clear()
                Password_field.send_keys(password)
                DRIVER.find_element_by_id("_submit").click()

                DRIVER.get(urljoin(f'{site}/translations/app/en/', domain))
                create_new()

                DRIVER.get(urljoin(f'{site}/translations/app/pl/', domain))
                create_new_pl()

                DRIVER.get(urljoin(f'{site}/translations/app/ref/', domain))
                create_new_ref()

                DRIVER.get(f'{site}/logout')
                break

            except selenium.common.exceptions.ElementClickInterceptedException:
                time.sleep(10)
                continue

            except selenium.common.exceptions.NoSuchElementException:
                DRIVER.get(f"{site}/logout")
                time.sleep(5)
                continue


translate()
if platform == "win32" or platform == "Windows":
    print("Script completed successfully")
    DRIVER.stop_client()
    DRIVER.quit()
    kill_driver()
    playsound(str(Paths.sound))

elif platform == "Linux" or platform == "Linux2":
    DRIVER.stop_client()
    DRIVER.quit()
    kill_driver()
    print("Script completed successfully")
