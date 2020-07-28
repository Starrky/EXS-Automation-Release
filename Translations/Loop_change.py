"""Keys and translations changed Prod/ Staging """
# !/usr/bin/python3.7.2
# coding=utf-8
import os
import sys
import time
from urllib.parse import urljoin
import openpyxl
import psutil
import selenium
import webdriver_manager.chrome
from playsound import playsound
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import platform

sys.path.append(
    os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)))
from configs import Accounts
from configs import Sites

# Driver settings
DRIVER = webdriver.Chrome(webdriver_manager.chrome.ChromeDriverManager().install())

platform = platform.system()
if platform == "win32" or platform == "Windows":
    from configs.Windows import Paths

elif platform == "Linux" or platform == "Linux2":
    from configs.Linux import Paths

websites_list = [Sites.Prod, Sites.Staging]
logins_list = [Accounts.Prod_user_1, Accounts.Staging_user_2]
passwords_list = [Accounts.Prod_Password_1, Accounts.Staging_Password_2]

language_code = ["en", "de", "sk", "tr", "ru", "cs", "nl", "hu", "ro", "zh", "pt", "es"]

# Grab text from launcher
currentdir = os.getcwd()
WB = openpyxl.load_workbook(f"{currentdir}/List.xlsx")
SHEET = WB.active
ROWNUM = SHEET.max_row

Keys = []
Domains = []
Old_pl = []
Old_en = []
New_pl = []
New_en = []

colDict = {1: Domains, 2: Keys, 3: Old_pl, 4: Old_en, 5: New_pl, 6: New_en}

for colNum, obj in colDict.items():
    for row in SHEET.iter_rows(min_row=2, min_col=colNum, max_col=colNum, max_row=ROWNUM):
        for cell in row:
            obj.append(cell.value)


def kill_driver():
    # End chromedriver and chrome ghost processes
    PROCNAME = "chromedriver.exe"

    for proc in psutil.process_iter():
        # check whether the process name matches
        if proc.name() == PROCNAME:
            try:
                proc.kill()

            except psutil.NoSuchProcess:
                break


def create_new_pl():
    for key, new_pl in zip(Keys, New_pl):
        """Create New polish translation for the key"""
        fa_plus = DRIVER.find_element_by_css_selector(".fa-plus")
        fa_plus.click()
        create_key = DRIVER.find_element_by_id("create-key")
        create_message = DRIVER.find_element_by_id("create-message")
        create_key.click()
        create_key.clear()
        create_key.send_keys(str(key))
        create_message.click()
        create_message.clear()
        create_message.send_keys(str(new_pl))
        create_message.send_keys(Keys.RETURN)


def create_new():
    for key, new_en in zip(Keys, New_en):
        """Create New english translation for the key(all languages, but polish and ref)"""
        fa_plus = DRIVER.find_element_by_css_selector(".fa-plus")
        fa_plus.click()
        create_key = DRIVER.find_element_by_id("create-key")
        create_message = DRIVER.find_element_by_id("create-message")
        create_key.click()
        create_key.clear()
        create_key.send_keys(str(key))
        create_message.click()
        create_message.clear()
        create_message.send_keys(str(new_en))
        create_message.send_keys(Keys.RETURN)


def create_new_ref():
    for key in Keys:
        """Create New REF translation for the key"""
        fa_plus = DRIVER.find_element_by_css_selector(".fa-plus")
        fa_plus.click()
        create_key = DRIVER.find_element_by_id("create-key")
        create_message = DRIVER.find_element_by_id("create-message")
        create_key.click()
        create_key.clear()
        create_key.send_keys(str(key))
        create_message.click()
        create_message.clear()
        create_message.send_keys(str(key))
        create_message.send_keys(Keys.RETURN)


def change():
    for key, old_pl, old_en, new_pl, new_en in zip(Keys, Old_pl, Old_en, New_pl, New_en):
        """Change the translation for the key if the value is equal to Old_eng or Old_pl(won't change
        translation that's different so already translated keys for other languages stays untouched"""
        try:
            value = DRIVER.find_element_by_css_selector(f"textarea[data-key='{key}']") \
                .get_attribute('value')
            field = DRIVER.find_element_by_css_selector(f"textarea[data-key='{key}']")

        except selenium.common.exceptions.NoSuchElementException:
            create_new()

        else:

            if value == old_pl:
                field.clear()
                field.send_keys(str(new_en))
                DRIVER.find_element_by_id(str(key)).click()

            elif value == old_en:
                field.clear()
                field.send_keys(str(new_en))
                DRIVER.find_element_by_id(str(key)).click()


def change_ref():
    for key in Keys:
        """Change the ref translation if key is not equal to key name(e.g if some1 did put pl/ en
        translation instead of app.key"""
        try:
            value = DRIVER.find_element_by_css_selector(f"textarea[data-key='{key}']") \
                .get_attribute('value')

            field = DRIVER.find_element_by_css_selector(f"textarea[data-key='{key}']")

        except selenium.common.exceptions.NoSuchElementException:
            create_new_ref()

        else:

            if value != key:
                field.clear()
                field.send_keys(str(key))
                DRIVER.find_element_by_id(str(key)).click()

            else:
                create_new_ref()


def change_pl():
    for key, old_pl, old_en, new_pl, new_en in zip(Keys, Old_pl, Old_en, New_pl, New_en):
        """Change polish translation for the key if it's translation is equal or different than Old_pl
        (so pretty much always changes it)"""
        try:
            field = DRIVER.find_element_by_css_selector(f"textarea[data-key='{key}']")
            value = field.get_attribute('value')
            # print(f"value is {value}")
        except selenium.common.exceptions.NoSuchElementException:
            create_new_pl()

        else:
            time.sleep(2)

            if value == old_pl:
                # print(f"{value} == {old_pl}")
                field.clear()
                field.send_keys(str(new_pl))
                DRIVER.find_element_by_id(str(key)).click()

            elif value != new_pl:
                # print(f"{value} != {new_pl}")
                field.clear()
                field.send_keys(str(new_pl))
                DRIVER.find_element_by_id(str(key)).click()

            else:
                print(f"{key} not found")
                create_new_pl()


def translate():
    """Actual Test execution code- takes domain from input in ENTRY4 and takes credentials from
    other file, then executes script on both staging and prod site"""
    DRIVER.start_client()
    for site, login, password, domain in zip(websites_list, logins_list, passwords_list, Domains):
        while True:
            try:
                DRIVER.get(site)
                Password_field = DRIVER.find_element_by_id("password")
                username = DRIVER.find_element_by_id("username")
                username.click()
                username.clear()
                username.send_keys(login)
                Password_field.click()
                Password_field.clear()
                Password_field.send_keys(password)
                DRIVER.find_element_by_id("_submit").click()
                while True:
                    try:
                        for lang in Domains:
                            DRIVER.get(urljoin(f'{site}/translations/app/{lang}/', domain))
                            change()
                            time.sleep(2)
                    finally:
                        break

                DRIVER.get(urljoin(f'{site}/translations/app/pl/', domain))
                change_pl()

                DRIVER.get(urljoin(f'{site}/translations/app/ref/', domain))
                change_ref()

            except selenium.common.exceptions.ElementClickInterceptedException:
                time.sleep(5)
                continue

            finally:
                break
        DRIVER.get(f"{site}/logout")


translate()

if platform == "win32" or platform == "Windows":
    print("Script completed successfully")
    DRIVER.stop_client()
    DRIVER.quit()
    kill_driver()
    playsound(str(Paths.sound))

elif platform == "Linux" or platform == "Linux2":
    DRIVER.stop_client()
    DRIVER.quit()
    kill_driver()
    print("Script completed successfully")
