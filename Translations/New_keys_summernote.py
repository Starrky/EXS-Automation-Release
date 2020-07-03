"""Keys and translations changed Prod/ Staging """
# !/usr/bin/python3.7.2
# coding=utf-8
import time
import sys
import os
from urllib.parse import urljoin
import openpyxl
import selenium
import webdriver_manager
from playsound import playsound
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import platform

sys.path.append(
    os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)))

from configs import Accounts
from configs import Sites

platform = platform.system()

if platform == "win32" or "Windows":
    from configs.Windows import Paths

elif platform == "Linux" or "Linux2":
    from configs.Linux import Paths


# Driver settings
DRIVER = webdriver.Chrome(webdriver_manager.chrome.ChromeDriverManager().install())

websites_list = [Sites.Prod, Sites.Staging]
logins_list = [Accounts.Prod_user_1, Accounts.Staging_user_2]
passwords_list = [Accounts.Prod_Password_1, Accounts.Staging_Password_2]
domains_name = ["", "staging."]
logout_list = [Sites.Staging_logout, Sites.Prod_logout]


# Grab text from launcher
WB = openpyxl.load_workbook(Paths.XLSX)
SHEET = WB.active
ROWNUM = SHEET.max_row

domain = SHEET.cell(row=ROWNUM, column=2).value
key = SHEET.cell(row=ROWNUM, column=3).value
new_pl = SHEET.cell(row=ROWNUM, column=4).value
new_en = SHEET.cell(row=ROWNUM, column=5).value
old_pl = SHEET.cell(row=ROWNUM, column=7).value
old_en = SHEET.cell(row=ROWNUM, column=8).value


def create_new_pl():
    """Create New polish translation for the key"""

    fa_plus = DRIVER.find_element_by_css_selector(".fa-plus")
    fa_plus.click()
    create_key = DRIVER.find_element_by_id("create-key")
    create_message = DRIVER.find_element_by_id("create-message")
    create_key.click()
    create_key.clear()
    create_key.send_keys(str(key))
    create_message.click()
    create_message.clear()
    create_message.send_keys(str(new_pl))
    create_message.send_keys(Keys.RETURN)


def create_new():
    """Create New english translation for the key(all languages, but polish and ref)"""

    fa_plus = DRIVER.find_element_by_css_selector(".fa-plus")
    fa_plus.click()
    create_key = DRIVER.find_element_by_id("create-key")
    add_editor = DRIVER.find_element_by_xpath("//*[@id='create-translation']/form/button[2]")
    create_key.click()
    create_key.clear()
    create_key.send_keys(str(key))
    add_editor.click()
    create_summernote = DRIVER.find_element_by_xpath(""
                                                     "//*[@id='create-translation']/form/div[2]/div/div[3]/div[2]/p")
    create_summernote.click()
    create_summernote.clear()
    create_summernote.send_keys(str(new_en))
    DRIVER.find_element_by_xpath("//*[@id='create-translation']/form/button[1]").click()


def create_new_ref():
    """Create New REF translation for the key"""

    fa_plus = DRIVER.find_element_by_css_selector(".fa-plus")
    fa_plus.click()
    create_key = DRIVER.find_element_by_id("create-key")
    create_message = DRIVER.find_element_by_id("create-message")
    create_key.click()
    create_key.clear()
    create_key.send_keys(str(key))
    create_message.click()
    create_message.clear()
    create_message.send_keys(str(key))
    create_message.send_keys(Keys.RETURN)


def translate():
    """Actual Test execution code- takes domain from input in ENTRY4 and takes credentials from
    other file, then executes script on both staging and prod site"""
    DRIVER.start_client()
    for site, login, password in zip(websites_list, logins_list, passwords_list):
        while True:
            try:
                DRIVER.get(site)
                Password_field = DRIVER.find_element_by_id("password")
                username = DRIVER.find_element_by_id("username")
                username.click()
                username.clear()
                username.send_keys(login)
                Password_field.click()
                Password_field.clear()
                Password_field.send_keys(password)
                DRIVER.find_element_by_id("_submit").click()

                DRIVER.get(urljoin(f'{site}/translations/app/en/', domain))
                create_new()

                DRIVER.get(urljoin(f'{site}/translations/app/pl/', domain))
                create_new_pl()

                DRIVER.get(urljoin(f'{site}/translations/app/ref/', domain))
                create_new_ref()

                DRIVER.get(f'{site}/logout')
                break

            except selenium.common.exceptions.ElementClickInterceptedException:
                time.sleep(10)
                continue

            except selenium.common.exceptions.NoSuchElementException:
                DRIVER.get(f"{site}/logout")
                time.sleep(5)
                continue
        break


translate()
playsound(str(Paths.sound))
